"""
TAX AI - Invoice Parser
Phân tích hóa đơn điện tử Việt Nam (.xml, .pdf)
Trích xuất MST bên bán / bên mua và địa chỉ.
Hỗ trợ chuẩn Nghị định 123/2020/NĐ-CP + Thông tư 78/2021/TT-BTC.
"""
import re
import logging
from dataclasses import dataclass, field
from difflib import SequenceMatcher
from typing import Optional
from datetime import datetime, timedelta
import xml.etree.ElementTree as ET

logger = logging.getLogger(__name__)

@dataclass
class InvoiceParty:
    mst: str = ""
    name: str = ""
    address: str = ""

@dataclass
class InvoiceData:
    seller: InvoiceParty = field(default_factory=InvoiceParty)
    buyer: InvoiceParty = field(default_factory=InvoiceParty)
    invoice_no: str = ""
    invoice_date: str = ""
    signing_time: str = ""
    source_format: str = ""
    parse_error: str = ""

@dataclass
class AddressMatch:
    invoice_address: str
    official_address: str
    similarity: float
    is_match: bool
    note: str = ""

@dataclass
class InvoiceVerifyResult:
    invoice: InvoiceData
    seller_info: object
    buyer_info: object
    seller_address_match: Optional[AddressMatch] = None
    buyer_address_match: Optional[AddressMatch] = None

_SELLER_MST_TAGS  = ["MST", "MSTNBan", "mstNBan", "taxCodeSeller", "taxCode"]
_SELLER_NAME_TAGS = ["TNBan", "tenNBan", "nameSeller", "tenCongTy", "Ten"]
_SELLER_ADDR_TAGS = ["DCNBan", "diaChiNBan", "addressSeller", "DChi", "diaChi"]
_BUYER_MST_TAGS   = ["MSTNMua", "mstNMua", "taxCodeBuyer", "MST"]
_BUYER_NAME_TAGS  = ["TNMua", "tenNMua", "nameBuyer", "tenKhachHang", "Ten"]
_BUYER_ADDR_TAGS  = ["DCNMua", "diaChiNMua", "addressBuyer", "DChi", "diaChi"]
_SELLER_SECTIONS  = ["NBan", "NguoiBan", "BenBan", "TTBan", "sellerInfo", "seller"]
_BUYER_SECTIONS   = ["NMua", "NguoiMua", "BenMua", "TTMua", "buyerInfo", "buyer"]

def _strip_ns(tag):
    return tag.split("}")[-1] if "}" in tag else tag

def _find_text(elem, *tag_names):
    for child in elem.iter():
        local = _strip_ns(child.tag)
        if local in tag_names and child.text and child.text.strip():
            return child.text.strip()
    return ""

def _find_section(root, *section_names):
    for elem in root.iter():
        if _strip_ns(elem.tag) in section_names:
            return elem
    return None

def _parse_xml(content: bytes) -> InvoiceData:
    data = InvoiceData(source_format="xml")
    try:
        root = ET.fromstring(content)
    except ET.ParseError as e:
        data.parse_error = f"XML không hợp lệ: {e}"
        return data
    seller_elem = _find_section(root, *_SELLER_SECTIONS)
    if seller_elem is not None:
        data.seller.mst     = _find_text(seller_elem, *_SELLER_MST_TAGS)
        data.seller.name    = _find_text(seller_elem, *_SELLER_NAME_TAGS)
        data.seller.address = _find_text(seller_elem, *_SELLER_ADDR_TAGS)
    buyer_elem = _find_section(root, *_BUYER_SECTIONS)
    if buyer_elem is not None:
        data.buyer.mst     = _find_text(buyer_elem, *_BUYER_MST_TAGS)
        data.buyer.name    = _find_text(buyer_elem, *_BUYER_NAME_TAGS)
        data.buyer.address = _find_text(buyer_elem, *_BUYER_ADDR_TAGS)
    if not data.seller.mst:
        for elem in root.iter():
            local = _strip_ns(elem.tag)
            if local in ("MSTNBan", "mstNBan") and elem.text and elem.text.strip():
                data.seller.mst = elem.text.strip(); break
        if not data.seller.mst:
            for elem in root.iter():
                if _strip_ns(elem.tag) == "MST" and elem.text and elem.text.strip():
                    data.seller.mst = elem.text.strip(); break
    if not data.buyer.mst:
        for elem in root.iter():
            local = _strip_ns(elem.tag)
            if local in ("MSTNMua", "mstNMua") and elem.text and elem.text.strip():
                data.buyer.mst = elem.text.strip(); break
    data.invoice_no   = _find_text(root, "SHDon", "soHoaDon", "invoiceNo", "KHSo", "so")
    data.invoice_date = _find_text(root, "NLap", "ngayLap", "invoiceDate", "ngayHoaDon")
    # Lấy thời gian ký số đầu tiên (bên bán)
    data.signing_time = _find_text(root, "SigningTime", "signingTime", "TGKy", "thoiGianKy")
    return data

_MST_RE       = re.compile(r"\b(\d{10}(?:-\d{3})?)\b")
_PHONE_FAX_RE = re.compile(r'^[\d\s\.\-\+\(\)\/]{7,}$')

_SELLER_KEYWORDS = [
    "người bán", "ben ban", "bên bán", "đơn vị bán", "nhà cung cấp",
    "công ty bán", "seller", "nban",
]
_BUYER_KEYWORDS = [
    "người mua", "ben mua", "bên mua", "đơn vị mua", "khách hàng",
    "công ty mua", "buyer", "nmua",
]
_ADDR_KEYWORDS = ["địa chỉ", "dia chi", "address", "đ/c:", "đc:", "đc "]
_FIELD_STOPS = [
    "mã số thuế", "mst", "điện thoại", "tel:", "phone", "fax", "email",
    "tài khoản", "ngân hàng", "họ tên", "hình thức", "ngày tháng",
    "số hóa đơn", "ký hiệu", "người mua", "người bán", "bên mua", "bên bán",
    "đơn vị mua", "đơn vị bán", "hàng hóa", "stt ", "tên hàng", "đvt",
]
_ADDR_INDICATOR = re.compile(
    r"(tầng\s*\d|tòa\s+nhà|toà\s+nhà|số\s+\d+|đường\s+\w|phường\s+\w|"
    r"quận\s+\d|huyện\s+\w|thành\s*phố|tp\.\s*hcm|tp\s+hà\s+nội|"
    r"p\.\s*\w|q\.\s*\d|khu\s+\w|lô\s+\w)",
    re.IGNORECASE | re.UNICODE,
)

def _clean(s):
    return s.replace("\xa0", " ").strip()

def _find_mst_near_keyword(lines, section_keywords):
    clean = [_clean(l) for l in lines]
    lower = [c.lower() for c in clean]
    for i, ll in enumerate(lower):
        if any(kw in ll for kw in section_keywords):
            m = _MST_RE.search(clean[i])
            if m:
                return m.group(1)
            for j in range(i + 1, min(len(clean), i + 21)):
                jl = lower[j]
                opposite = _BUYER_KEYWORDS if section_keywords == _SELLER_KEYWORDS else _SELLER_KEYWORDS
                if any(kw in jl for kw in opposite):
                    break
                m = _MST_RE.search(clean[j])
                if m:
                    return m.group(1)
    return ""

def _find_address_near_keyword(lines, section_keywords):
    clean = [_clean(l) for l in lines]
    lower = [c.lower() for c in clean]
    section_line = 0
    for i, ll in enumerate(lower):
        if any(kw in ll for kw in section_keywords):
            section_line = i
            break
    addr_line = -1
    for i in range(section_line, min(len(clean), section_line + 15)):
        if any(kw in lower[i] for kw in _ADDR_KEYWORDS):
            addr_line = i
            break
    if addr_line == -1:
        return ""
    line = clean[addr_line]
    colon = line.find(":")
    value = line[colon + 1:].strip() if colon != -1 else line.strip()
    if value:
        for i in range(addr_line + 1, min(len(clean), addr_line + 6)):
            s = clean[i]
            ll = lower[i]
            if not s:
                break
            if any(kw in ll for kw in _FIELD_STOPS):
                break
            colon_pos = s.find(":")
            if 0 < colon_pos < 30:
                break
            value = (value + ", " + s).strip(", ")
    else:
        opposite = _BUYER_KEYWORDS if section_keywords is _SELLER_KEYWORDS else _SELLER_KEYWORDS
        for i in range(addr_line + 1, min(len(clean), addr_line + 25)):
            s = clean[i]
            ll = lower[i]
            if not s:
                continue
            if any(kw in ll for kw in opposite):
                break
            colon_pos = s.find(":")
            if 0 < colon_pos < 30:
                continue
            if re.match(r'^[\d\s\.\-,]+$', s):
                continue
            if _ADDR_INDICATOR.search(ll):
                value = s
                if i + 1 < len(clean):
                    ns = clean[i + 1]
                    nl = lower[i + 1] if i + 1 < len(lower) else ""
                    nc = ns.find(":")
                    if ns and not (0 < nc < 30) and not re.match(r'^[\d\s\.\-,]+$', ns):
                        if not any(kw in nl for kw in _FIELD_STOPS):
                            value = (value + ", " + ns).strip(", ")
                break
    return value

def _find_field_value(lines, field_keywords, start=0, end=-1):
    clean = [_clean(l) for l in lines]
    lower = [c.lower() for c in clean]
    if end < 0:
        end = len(lines)
    for i in range(start, min(end, len(clean))):
        if any(kw in lower[i] for kw in field_keywords):
            line = clean[i]
            colon = line.find(":")
            value = line[colon + 1:].strip() if colon != -1 else line.strip()
            if value:
                return value
            if i + 1 < end:
                nxt = clean[i + 1].strip()
                colon_pos = nxt.find(":")
                if nxt and not (0 < colon_pos < 30):
                    return nxt
    return ""

def _find_section_bounds(lines, section_keywords, stop_keywords=None, use_last=False):
    lower = [_clean(l).lower() for l in lines]
    start = -1
    if use_last:
        for i in range(len(lower) - 1, -1, -1):
            if any(kw in lower[i] for kw in section_keywords):
                start = i; break
    else:
        for i, ll in enumerate(lower):
            if any(kw in ll for kw in section_keywords):
                start = i; break
    if start == -1:
        return -1, len(lines)
    end = len(lines)
    if stop_keywords:
        for i in range(start + 1, len(lower)):
            if any(kw in lower[i] for kw in stop_keywords):
                end = i; break
    return start, end

def _parse_pdf(content: bytes) -> InvoiceData:
    data = InvoiceData(source_format="pdf")
    try:
        import fitz
    except ImportError:
        data.parse_error = "Thiếu thư viện pymupdf. Chạy: pip install pymupdf"
        return data
    try:
        doc = fitz.open(stream=content, filetype="pdf")
        full_text = "\n".join(page.get_text() for page in doc)
        doc.close()
    except Exception as e:
        data.parse_error = f"Không đọc được PDF: {e}"
        return data
    if not full_text.strip():
        data.parse_error = "PDF không chứa text (có thể là ảnh scan)"
        return data

    lines = full_text.splitlines()
    clean = [_clean(l) for l in lines]
    lower = [c.lower() for c in clean]

    _SELLER_SECTION_KW = ["đơn vị bán hàng", "bên bán", "người bán hàng", "seller"]
    _BUYER_SECTION_KW  = [
        "họ tên người mua", "tên đơn vị", "bên mua", "đơn vị mua",
        "người mua hàng", "buyer",
    ]

    s_start, s_end = _find_section_bounds(lines, _SELLER_SECTION_KW,
                                          stop_keywords=_BUYER_SECTION_KW)
    b_start, b_end = _find_section_bounds(lines, _BUYER_SECTION_KW, use_last=False)
    if 0 <= b_start < s_end and s_end < len(lines):
        lower_full = [_clean(l).lower() for l in lines]
        tmp = -1
        for i in range(s_end, len(lower_full)):
            if any(kw in lower_full[i] for kw in _BUYER_SECTION_KW):
                tmp = i; break
        if tmp >= 0:
            b_start, b_end = tmp, len(lines)

    if s_start >= 0:
        data.seller.mst = (_find_mst_near_keyword(lines[s_start:s_end], _SELLER_KEYWORDS)
                           or _find_mst_near_keyword(lines[s_start:s_end], ["mã số thuế", "mst"]))
    if b_start >= 0:
        data.buyer.mst = (_find_mst_near_keyword(lines[b_start:b_end], _BUYER_KEYWORDS)
                          or _find_mst_near_keyword(lines[b_start:b_end], ["mã số thuế", "mst"]))

    if not data.seller.mst or not data.buyer.mst:
        all_msts = list(dict.fromkeys(_MST_RE.findall("\n".join(_clean(l) for l in lines))))
        if len(all_msts) >= 1 and not data.seller.mst:
            data.seller.mst = all_msts[0]
        if len(all_msts) >= 2 and not data.buyer.mst:
            data.buyer.mst = all_msts[1]

    if s_start >= 0:
        for i in range(s_start, min(s_end, s_start + 25)):
            s = clean[i]; ll = lower[i]
            if not s: continue
            colon_pos = s.find(":")
            if 0 < colon_pos < 35:
                val = s[colon_pos + 1:].strip()
                if val and not _MST_RE.fullmatch(val.replace("-", "")) and not _PHONE_FAX_RE.match(val):
                    data.seller.name = val; break
                continue
            if _MST_RE.fullmatch(s.replace("-", "")) or re.match(r'^[\d\s\.\-,\.\(\)]+$', s) or len(s) < 5:
                continue
            if len(s) >= 10 and not any(kw in ll for kw in _FIELD_STOPS):
                data.seller.name = s; break

    if b_start >= 0:
        data.buyer.name = _find_field_value(lines, ["tên đơn vị", "tên công ty", "company"],
                                            start=b_start, end=b_end)
        if not data.buyer.name:
            for i in range(b_start, min(b_end, b_start + 15)):
                s = clean[i]; ll = lower[i]
                if not s or len(s) < 5: continue
                colon_pos = s.find(":")
                if 0 < colon_pos < 35:
                    val = s[colon_pos + 1:].strip()
                    if val and not _MST_RE.fullmatch(val.replace("-", "")) and not _PHONE_FAX_RE.match(val):
                        data.buyer.name = val; break
                    continue

    if s_start >= 0:
        data.seller.address = _find_address_near_keyword(
            lines[s_start:s_end], _SELLER_KEYWORDS + ["đơn vị bán hàng", "địa chỉ"])
        if not data.seller.address:
            data.seller.address = _find_address_near_keyword(
                lines[s_start:s_end], ["địa chỉ", "address"])

    if b_start >= 0:
        data.buyer.address = _find_address_near_keyword(
            lines[b_start:b_end], _BUYER_KEYWORDS + ["họ tên người mua", "địa chỉ"])
        if not data.buyer.address:
            data.buyer.address = _find_address_near_keyword(
                lines[b_start:b_end], ["địa chỉ", "address"])

    data.seller.address = re.sub(r",\s*,", ",", data.seller.address).strip(", ")
    data.buyer.address  = re.sub(r",\s*,", ",", data.buyer.address).strip(", ")

    inv_m = re.search(r"(?:số|ký hiệu)[:\s]*([A-Z0-9/\-]+)", full_text, re.IGNORECASE)
    if inv_m:
        data.invoice_no = inv_m.group(1).strip()

    return data

def parse_invoice(content: bytes, filename: str) -> InvoiceData:
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if ext == "xml":
        return _parse_xml(content)
    elif ext == "pdf":
        return _parse_pdf(content)
    else:
        if content.lstrip()[:5] in (b"<?xml", b"<HDDT", b"<inv:"):
            return _parse_xml(content)
        if content[:4] == b"%PDF":
            return _parse_pdf(content)
        data = InvoiceData()
        data.parse_error = f"Định dạng không hỗ trợ: {ext or 'unknown'}. Chỉ hỗ trợ .xml và .pdf"
        return data

def _normalize_address(addr: str) -> str:
    addr = addr.lower().strip()
    addr = re.sub(r"\s+", " ", addr)
    addr = re.sub(r"[,\.]+", ",", addr)
    return addr

def _raw_equal(a: str, b: str) -> bool:
    """So sánh địa chỉ ở mức raw (chỉ chuẩn hóa khoảng trắng và hoa/thường)."""
    return re.sub(r"\s+", " ", a.lower().strip()) == re.sub(r"\s+", " ", b.lower().strip())

def compare_address(invoice_addr: str, official_addr: str, threshold: float = 0.65) -> AddressMatch:
    if not invoice_addr or not official_addr:
        note = "Hóa đơn không có địa chỉ" if not invoice_addr else "Không tra cứu được địa chỉ từ GDT"
        return AddressMatch(invoice_address=invoice_addr, official_address=official_addr,
                            similarity=0.0, is_match=False, note=note)
    # Chỉ coi là khớp khi chuỗi thực sự giống nhau (không normalize mạnh)
    if _raw_equal(invoice_addr, official_addr):
        return AddressMatch(invoice_address=invoice_addr, official_address=official_addr,
                            similarity=1.0, is_match=True, note="Khớp (100%)")
    a = _normalize_address(invoice_addr)
    b = _normalize_address(official_addr)
    sim = SequenceMatcher(None, a, b).ratio()
    a_tokens = set(a.replace(",", " ").split()) - {"và","the","a","in","of","tại","số","-",",","."}
    b_tokens = set(b.replace(",", " ").split()) - {"và","the","a","in","of","tại","số","-",",","."}
    token_overlap = 0.0
    if a_tokens and b_tokens:
        common = a_tokens & b_tokens
        token_overlap = len(common) / max(len(a_tokens), len(b_tokens))
    combined = 0.6 * sim + 0.4 * token_overlap
    # Không dùng threshold để báo "Khớp" — chỉ raw equal mới là Khớp
    if combined >= 0.75:
        note = f"Gần khớp ({combined:.0%}) — cần kiểm tra"
    elif combined >= 0.4:
        note = f"Có thể trùng ({combined:.0%}) — cần kiểm tra thủ công"
    else:
        note = f"Không khớp ({combined:.0%})"
    return AddressMatch(invoice_address=invoice_addr, official_address=official_addr,
                        similarity=combined, is_match=False, note=note)

def _red(text: str) -> str:
    return f"[[R]]{text}[[/R]]"


def _name_differs(inv_name: str, gdt_name: str, threshold: float = 0.55) -> bool:
    """Trả về True nếu tên trên hóa đơn khác đáng kể so với tên GDT."""
    if not inv_name or not gdt_name:
        return False
    a = inv_name.lower().strip()
    b = gdt_name.lower().strip()
    if a == b:
        return False
    sim = SequenceMatcher(None, a, b).ratio()
    a_tok = set(a.split()) - {"cong", "ty", "co", "phan", "tnhh", "mtv", "ltd", "jsc"}
    b_tok = set(b.split()) - {"cong", "ty", "co", "phan", "tnhh", "mtv", "ltd", "jsc"}
    tok_overlap = len(a_tok & b_tok) / max(len(a_tok), len(b_tok), 1)
    combined = 0.6 * sim + 0.4 * tok_overlap
    return combined < threshold


def _addr_block(match):
    if not match:
        return f"  📍 Địa chỉ: {_red('Không có dữ liệu')}"
    inv_addr = match.invoice_address or "(trống)"
    off_addr = match.official_address or "(trống)"
    if match.is_match:
        # Khớp 100% — chỉ hiện 1 dòng, không cần so sánh
        lines = [f"  📍 Địa chỉ: {inv_addr}",
                 f"  ✅ Đối chiếu: {match.note}"]
    elif match.similarity >= 0.75:
        # Gần khớp — bôi đỏ cả 2 dòng, cảnh báo ⚠️
        lines = [f"  📍 Địa chỉ HĐ : {_red(inv_addr)}",
                 f"  📍 Địa chỉ GDT: {_red(off_addr)}",
                 f"  ⚠️ {_red('Đối chiếu: ' + match.note)}"]
    else:
        # Không khớp — bôi đỏ cả 2 dòng, lỗi ❌
        lines = [f"  📍 Địa chỉ HĐ : {_red(inv_addr)}",
                 f"  📍 Địa chỉ GDT: {_red(off_addr)}",
                 f"  ❌ {_red('Đối chiếu: ' + match.note)}"]
    return "\n".join(lines)


def _check_signing_delay(invoice_date: str, signing_time: str, max_hours: int = 24):
    """Trả về (delay_hours, warning_str) nếu ký số trễ hơn max_hours sau ngày lập HĐ."""
    if not invoice_date or not signing_time:
        return None, None
    try:
        # invoice_date: "2026-05-17" hoặc "2026-05-17T..."
        if "T" in invoice_date:
            inv_dt = datetime.fromisoformat(invoice_date)
        else:
            inv_dt = datetime.strptime(invoice_date, "%Y-%m-%d")
        # signing_time: "2026-05-18T13:55:41" hoặc có timezone
        st = signing_time.split("+")[0].split("Z")[0]
        sign_dt = datetime.fromisoformat(st)
        delta = sign_dt - inv_dt
        hours = delta.total_seconds() / 3600
        if hours > max_hours:
            d = int(hours // 24)
            h = int(hours % 24)
            delay_str = f"{d} ngày {h} giờ" if d > 0 else f"{h} giờ"
            return hours, delay_str
        return hours, None
    except Exception:
        return None, None


def format_verify_result(result: InvoiceVerifyResult) -> str:
    inv   = result.invoice
    lines = ["📄 *Kết quả kiểm tra hóa đơn điện tử*"]
    if inv.invoice_no:
        lines.append(f"🔖 Số HĐ: `{inv.invoice_no}`")
    if inv.invoice_date:
        lines.append(f"📅 Ngày lập: {inv.invoice_date}")
    if inv.signing_time:
        lines.append(f"🖊️ Ngày ký: {inv.signing_time}")
    lines.append(f"📁 Định dạng: {inv.source_format.upper()}")
    if inv.parse_error:
        lines.append(f"\n{_red('❌ Lỗi đọc file: ' + inv.parse_error)}")
        return "\n".join(lines)

    lines.append("\n─────────────────────")
    lines.append("🏭 *BÊN BÁN*")
    if inv.seller.mst:
        lines.append(f"  🔢 MST: `{inv.seller.mst}`")
    else:
        lines.append(f"  {_red('⚠️ Không tìm được MST bên bán')}")
    if result.seller_info and not result.seller_info.error:
        si = result.seller_info
        # So sánh tên công ty
        if inv.seller.name:
            inv_n = inv.seller.name.strip()
            gdt_n = si.name.strip()
            if inv_n.lower() == gdt_n.lower():
                lines.append(f"  🏢 Tên: {gdt_n}")
            elif _name_differs(inv_n, gdt_n):
                lines.append(f"  🏢 Tên HĐ : {_red(inv_n)}")
                lines.append(f"  🏢 Tên GDT: {gdt_n}")
            else:
                lines.append(f"  🏢 Tên HĐ : {inv_n}")
                lines.append(f"  🏢 Tên GDT: {gdt_n}")
                lines.append(f"  ⚠️ {_red('Tên gần khớp — kiểm tra lại')}")
        else:
            lines.append(f"  🏢 Tên (GDT): {si.name}")
        if si.is_active:
            lines.append(f"  ✅ Trạng thái: {si.status}")
        else:
            lines.append(f"  {_red('❌ Trạng thái: ' + si.status)}")
        lines.append(_addr_block(result.seller_address_match))
    elif inv.seller.mst:
        lines.append(f"  {_red('❌ Không tra cứu được MST ' + inv.seller.mst)}")

    lines.append("\n─────────────────────")
    lines.append("🛒 *BÊN MUA*")
    if inv.buyer.mst:
        lines.append(f"  🔢 MST: `{inv.buyer.mst}`")
    else:
        lines.append(f"  {_red('⚠️ Không tìm được MST bên mua')}")
    if result.buyer_info and not result.buyer_info.error:
        bi = result.buyer_info
        # So sánh tên công ty
        if inv.buyer.name:
            inv_n = inv.buyer.name.strip()
            gdt_n = bi.name.strip()
            if inv_n.lower() == gdt_n.lower():
                lines.append(f"  🏢 Tên: {gdt_n}")
            elif _name_differs(inv_n, gdt_n):
                lines.append(f"  🏢 Tên HĐ : {_red(inv_n)}")
                lines.append(f"  🏢 Tên GDT: {gdt_n}")
            else:
                lines.append(f"  🏢 Tên HĐ : {inv_n}")
                lines.append(f"  🏢 Tên GDT: {gdt_n}")
                lines.append(f"  ⚠️ {_red('Tên gần khớp — kiểm tra lại')}")
        else:
            lines.append(f"  🏢 Tên (GDT): {bi.name}")
        if bi.is_active:
            lines.append(f"  ✅ Trạng thái: {bi.status}")
        else:
            lines.append(f"  {_red('❌ Trạng thái: ' + bi.status)}")
        lines.append(_addr_block(result.buyer_address_match))
    elif inv.buyer.mst:
        lines.append(f"  {_red('❌ Không tra cứu được MST ' + inv.buyer.mst)}")

    lines.append("\n─────────────────────")
    issues = []
    if result.seller_info and result.seller_info.error:
        issues.append("MST bên bán không hợp lệ")
    if result.seller_info and not result.seller_info.is_active:
        issues.append("Bên bán đã ngừng hoạt động")
    if inv.seller.name and result.seller_info and not result.seller_info.error:
        if _name_differs(inv.seller.name, result.seller_info.name):
            issues.append("Tên bên bán trên HĐ không khớp GDT")
    if result.seller_address_match and not result.seller_address_match.is_match:
        issues.append("Địa chỉ bên bán không khớp GDT")
    if result.buyer_info and result.buyer_info.error:
        issues.append("MST bên mua không hợp lệ")
    if result.buyer_info and not result.buyer_info.is_active:
        issues.append("Bên mua đã ngừng hoạt động")
    if inv.buyer.name and result.buyer_info and not result.buyer_info.error:
        if _name_differs(inv.buyer.name, result.buyer_info.name):
            issues.append("Tên bên mua trên HĐ không khớp GDT")
    if result.buyer_address_match and not result.buyer_address_match.is_match:
        issues.append("Địa chỉ bên mua không khớp GDT")
    _, sign_delay = _check_signing_delay(inv.invoice_date, inv.signing_time)
    if sign_delay:
        issues.append(f"Ký số trễ {sign_delay} sau ngày lập hóa đơn (>24h)")

    if issues:
        lines.append(_red("⚠️ CẦN LƯU Ý:"))
        for issue in issues:
            lines.append(_red(f"  • {issue}"))
    else:
        lines.append("✅ Hóa đơn không có dấu hiệu bất thường")
    return "\n".join(lines)


def verify_invoice(content: bytes, filename: str) -> InvoiceVerifyResult:
    from mst_lookup import lookup_mst, is_valid_mst
    invoice = parse_invoice(content, filename)
    result  = InvoiceVerifyResult(invoice=invoice, seller_info=None, buyer_info=None)
    if invoice.parse_error:
        return result
    if invoice.seller.mst and is_valid_mst(invoice.seller.mst):
        result.seller_info = lookup_mst(invoice.seller.mst)
        if result.seller_info and not result.seller_info.error:
            result.seller_address_match = compare_address(
                invoice.seller.address, result.seller_info.address)
    if invoice.buyer.mst and is_valid_mst(invoice.buyer.mst):
        result.buyer_info = lookup_mst(invoice.buyer.mst)
        if result.buyer_info and not result.buyer_info.error:
            result.buyer_address_match = compare_address(
                invoice.buyer.address, result.buyer_info.address)
    return result
# ─────────────────────────────────────────────
# Batch processing helpers
# ─────────────────────────────────────────────

def _extract_batch_row(filename, result):
    """Trich xuat du lieu tu InvoiceVerifyResult thanh dict cho batch Excel."""
    inv = result.invoice
    row = {
        "filename":      filename,
        "invoice_no":    inv.invoice_no,
        "invoice_date":  inv.invoice_date,
        "signing_time":  inv.signing_time,
        "parse_error":   inv.parse_error,
        "seller_name_inv": inv.seller.name,
        "seller_mst":      inv.seller.mst,
        "seller_name_gdt": "",
        "seller_status":   "",
        "seller_is_active": None,
        "seller_addr_note": "",
        "buyer_name_inv":  inv.buyer.name,
        "buyer_mst":       inv.buyer.mst,
        "buyer_name_gdt":  "",
        "buyer_status":    "",
        "buyer_is_active": None,
        "buyer_addr_note": "",
        "signing_delay_str":   "",
        "signing_delay_hours": None,
    }
    if result.seller_info:
        if not result.seller_info.error:
            si = result.seller_info
            row["seller_name_gdt"]  = si.name
            row["seller_status"]    = si.status
            row["seller_is_active"] = si.is_active
        else:
            row["seller_status"]    = "Khong tra cuu duoc"
            row["seller_is_active"] = False
    elif inv.seller.mst:
        row["seller_status"]    = "Khong tra cuu duoc"
        row["seller_is_active"] = False
    if result.buyer_info:
        if not result.buyer_info.error:
            bi = result.buyer_info
            row["buyer_name_gdt"]  = bi.name
            row["buyer_status"]    = bi.status
            row["buyer_is_active"] = bi.is_active
        else:
            row["buyer_status"]    = "Khong tra cuu duoc"
            row["buyer_is_active"] = False
    elif inv.buyer.mst:
        row["buyer_status"]    = "Khong tra cuu duoc"
        row["buyer_is_active"] = False
    if result.seller_address_match:
        row["seller_addr_note"] = result.seller_address_match.note
    if result.buyer_address_match:
        row["buyer_addr_note"] = result.buyer_address_match.note
    hours, delay_str = _check_signing_delay(inv.invoice_date, inv.signing_time)
    row["signing_delay_hours"] = hours
    row["signing_delay_str"]   = delay_str or ""
    return row


def batch_to_excel(rows):
    """Tao file Excel 2 sheet tu list dict (output cua _extract_batch_row)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("Thieu thu vien openpyxl. Chay: pip install openpyxl")
    from datetime import datetime as _dt
    from io import BytesIO

    wb = openpyxl.Workbook()
    BLUE = "1565C0"; RED = "C62828"; RED_F = "FFEBEE"; GREEN = "2E7D32"
    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill  = PatternFill("solid", fgColor=BLUE)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    red_font  = Font(bold=True, color=RED)
    sec_font  = Font(bold=True, size=11, color=BLUE)
    wrap_align = Alignment(vertical="center", wrap_text=True)

    def _side():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    def hdr_row(ws, r, cols):
        for c, txt in enumerate(cols, 1):
            cell = ws.cell(row=r, column=c, value=txt)
            cell.font = hdr_font; cell.fill = hdr_fill
            cell.alignment = hdr_align; cell.border = _side()
        ws.row_dimensions[r].height = 28

    def data_row(ws, r, vals, highlight=False):
        fill = PatternFill("solid", fgColor=RED_F) if highlight else None
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = wrap_align; cell.border = _side()
            if highlight:
                cell.fill = fill; cell.font = red_font

    # Sheet 2: Chi tiet
    ws2 = wb.active
    ws2.title = "Chi tiet"
    COLS2 = ["STT", "Ten File", "So HD",
             "Ben ban", "MST ban", "TT ban",
             "Ben mua", "MST mua", "TT mua",
             "DC ben ban", "DC ben mua", "Chenh lech ngay ky"]
    hdr_row(ws2, 1, COLS2)
    WIDTHS2 = [5, 28, 14, 30, 14, 26, 30, 14, 26, 32, 32, 22]
    for i, w in enumerate(WIDTHS2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    for idx, row in enumerate(rows, 1):
        s_active = row.get("seller_is_active")
        b_active = row.get("buyer_is_active")
        dh       = row.get("signing_delay_hours")
        flagged  = (s_active is False) or (b_active is False) or (dh is not None and dh > 24)
        s_name = row.get("seller_name_gdt") or row.get("seller_name_inv") or ""
        b_name = row.get("buyer_name_gdt")  or row.get("buyer_name_inv")  or ""
        if dh is None:
            delay_disp = ""
        elif dh <= 24:
            delay_disp = f"{dh:.1f}h (OK)"
        else:
            delay_disp = row.get("signing_delay_str") or f"{dh:.1f}h"
        data_row(ws2, idx + 1, [
            idx,
            row.get("filename", ""),
            row.get("invoice_no", ""),
            s_name, row.get("seller_mst", ""), row.get("seller_status", ""),
            b_name, row.get("buyer_mst", ""),  row.get("buyer_status", ""),
            row.get("seller_addr_note", ""),
            row.get("buyer_addr_note", ""),
            delay_disp,
        ], highlight=flagged)

    # Sheet 1: Tong hop
    ws1 = wb.create_sheet("Tong hop", 0)
    ws1.merge_cells("A1:G1")
    c = ws1["A1"]
    c.value = "BAO CAO KIEM TRA HOA DON DIEN TU HANG LOAT"
    c.font = Font(bold=True, size=14, color=BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 32
    ws1.merge_cells("A2:G2")
    ws1["A2"].value = "Ngay tao: " + _dt.now().strftime("%d/%m/%Y %H:%M")
    ws1["A2"].alignment = Alignment(horizontal="center")

    total        = len(rows)
    inact_seller = [r for r in rows if r.get("seller_is_active") is False]
    inact_buyer  = [r for r in rows if r.get("buyer_is_active")  is False]
    delayed      = [r for r in rows if (r.get("signing_delay_hours") or 0) > 24]

    r = 4
    ws1[f"A{r}"].value = "I. TONG HOP"; ws1[f"A{r}"].font = sec_font; r += 1
    for lbl, val, warn in [
        ("Tong so hoa don kiem tra:", total, False),
        ("Hoa don ben ban khong hoat dong:", len(inact_seller), bool(inact_seller)),
        ("Hoa don ben mua khong hoat dong:", len(inact_buyer),  bool(inact_buyer)),
        ("Hoa don ky so tre >24h:",          len(delayed),      bool(delayed)),
    ]:
        ws1[f"A{r}"].value = lbl
        ws1[f"B{r}"].value = val
        if warn: ws1[f"B{r}"].font = red_font
        r += 1

    ws1.column_dimensions["A"].width = 40
    ws1.column_dimensions["B"].width = 12
    for col, w in [("C",14),("D",12),("E",35),("F",15),("G",30)]:
        ws1.column_dimensions[col].width = w

    r += 1
    ws1[f"A{r}"].value = "II. HOA DON CO MST KHONG HOAT DONG"; ws1[f"A{r}"].font = sec_font; r += 1
    inactive_all = []
    for row in rows:
        if row.get("seller_is_active") is False:
            inactive_all.append(dict(row, _side="Ben ban",
                _mst=row.get("seller_mst",""),
                _name=row.get("seller_name_gdt") or row.get("seller_name_inv",""),
                _status=row.get("seller_status","")))
        if row.get("buyer_is_active") is False:
            inactive_all.append(dict(row, _side="Ben mua",
                _mst=row.get("buyer_mst",""),
                _name=row.get("buyer_name_gdt") or row.get("buyer_name_inv",""),
                _status=row.get("buyer_status","")))
    if inactive_all:
        hdr_row(ws1, r, ["STT","Ten File","So HD","Ben","Ten DN","MST","Trang thai"]); r += 1
        for i, ir in enumerate(inactive_all, 1):
            data_row(ws1, r, [i, ir.get("filename",""), ir.get("invoice_no",""),
                               ir["_side"], ir["_name"], ir["_mst"], ir["_status"]], highlight=True)
            r += 1
    else:
        ws1[f"A{r}"].value = "Khong co hoa don nao co MST khong hoat dong"
        ws1[f"A{r}"].font = Font(color=GREEN); r += 1

    r += 1
    ws1[f"A{r}"].value = "III. HOA DON KY SO TRE >24H"; ws1[f"A{r}"].font = sec_font; r += 1
    if delayed:
        hdr_row(ws1, r, ["STT","Ten File","So HD","Ngay lap","Ngay ky","Chenh lech"]); r += 1
        for i, dr in enumerate(delayed, 1):
            data_row(ws1, r, [i, dr.get("filename",""), dr.get("invoice_no",""),
                               dr.get("invoice_date",""), dr.get("signing_time",""),
                               dr.get("signing_delay_str","")], highlight=True)
            r += 1
    else:
        ws1[f"A{r}"].value = "Khong co hoa don nao ky so tre >24h"
        ws1[f"A{r}"].font = Font(color=GREEN)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
